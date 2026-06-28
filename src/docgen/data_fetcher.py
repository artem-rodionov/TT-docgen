import requests
from openpyxl import load_workbook

class DataFetcher:
    def __init__(self, token: str, path: str) -> None:
        self.token = token
        self.table = load_workbook(path, data_only=True)

    def get_projects(self):
        '''Получение списка всех проектов.'''
        url = f"https://uchet.type-tech.ru/api/projects?token={self.token}"
        response = requests.get(url)
        response.raise_for_status()
        return response.json()["projects"]
    
    def get_project_info(self, id):
        '''Получение информации о проекте.'''
        url = f"https://uchet.type-tech.ru/api/project/{id}/authors-summary?token={self.token}"
        response = requests.get(url)
        response.raise_for_status()
        return response.json()
    
    def get_workers_mapping(self):
        '''Получение соответствия ФИО и имен из БД.'''
        sheet = self.table.worksheets[1]
        names = {}
        for row in sheet.iter_rows(values_only=True):
            i = 0
            while i < len(row):
                col = row[i]
                if col is not None:
                    if row[i+1] is None:
                        raise ValueError(col)
                    names[col] = row[i+1]
                    i += 2
                    continue
                i += 1
        return names

    def get_workers_data(self):
        '''Получение данных о работниках.'''
        sheet = self.table.worksheets[0]
        workers = {}
        for row in sheet.iter_rows(values_only=True):
            if row[2] is None:
                continue
            other_cols = [row[i] for i in range(15) if i != 2]

            workers[row[2]] = other_cols

        return workers